import os
import requests
import pandas as pd

MASTER_FILE = "유월의보리_product_master.xlsx"
ERP_FILE = "erp_sales.xlsx"

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")

TELEGRAM_CHAT_IDS = [
    "1490548765",   # 도현
]

def to_number(value):
    if pd.isna(value):
        return 0
    return float(str(value).replace(",", "").strip())


STORE_ORDER = {
    "양재": 0,
    "신흥": 1,
    "본점": 2,
    "신내": 3,
}


def normalize_store_name(value):
    if pd.isna(value):
        return value

    name = str(value).strip()

    if "양재" in name:
        return "양재"
    if "신흥" in name:
        return "신흥"
    if "본점" in name:
        return "본점"
    if "신내" in name:
        return "신내"

    return name


def run_monthly_logistics_report():

    master = pd.read_excel(
        MASTER_FILE,
        sheet_name="PRODUCT_MASTER"
    )

    erp = pd.read_excel(
        ERP_FILE,
        sheet_name="판매현황",
        header=1
    )

    erp.columns = erp.columns.str.strip()
    master.columns = master.columns.str.strip()

    erp = erp[
        erp["품목코드"].notna()
    ]

    erp["거래처명"] = erp["거래처명"].apply(normalize_store_name)

    master["품목코드"] = master["품목코드"].astype(str).str.strip()
    erp["품목코드"] = erp["품목코드"].astype(str).str.strip()

    master = master[
        master["사용여부"].astype(str).str.upper() == "Y"
    ]

    cost_map = dict(
        zip(
            master["품목코드"],
            master["제조원가(입력)"]
        )
    )

    erp["수량"] = erp["수량"].apply(to_number)
    erp["공급가액"] = erp["공급가액"].apply(to_number)
    erp["부가세"] = erp["부가세"].apply(to_number)
    erp["합계"] = erp["합계"].apply(to_number)

    erp["제조원가_단가"] = erp["품목코드"].map(cost_map)
    missing = erp[erp["제조원가_단가"].isna()][["품목코드", "품목명(규격)"]].drop_duplicates()

    erp["총제조원가"] = erp["수량"] * erp["제조원가_단가"].fillna(0)
    erp["물류이익"] = erp["공급가액"] - erp["총제조원가"]

    store_report = (
        erp.groupby("거래처명")
        .agg(
            공급가액=("공급가액","sum"),
            부가세=("부가세","sum"),
            청구합계=("합계","sum"),
            총제조원가=("총제조원가","sum"),
            물류이익=("물류이익","sum"),
        )
        .reset_index()
    )

    store_report["_sort"] = store_report["거래처명"].map(STORE_ORDER).fillna(99)
    store_report = (
        store_report
        .sort_values(["_sort", "거래처명"])
        .drop(columns=["_sort"])
        .reset_index(drop=True)
    )
    
    store_report["물류이익률"] = (
    store_report["물류이익"] /
    store_report["공급가액"].replace(0, pd.NA)
)
    
    sku_report = (
        erp.groupby(["품목코드", "품목명(규격)"])
        .agg(
            수량=("수량", "sum"),
            공급가액=("공급가액", "sum"),
            총제조원가=("총제조원가", "sum"),
            물류이익=("물류이익", "sum"),
        )
        .reset_index()
    )

    sku_report["물류이익률"] = sku_report["물류이익"] / sku_report["공급가액"].replace(0, pd.NA)

    lines = []

    total_supply = store_report["공급가액"].sum()
    total_vat = store_report["부가세"].sum()
    total_invoice = store_report["청구합계"].sum()
    total_cost = store_report["총제조원가"].sum()
    total_profit = store_report["물류이익"].sum()
    total_margin = total_profit / total_supply if total_supply != 0 else 0

    lines.append(f"""
[월말 물류이익 리포트]

TOTAL
공급가액: {total_supply:,.0f}원
부가세: {total_vat:,.0f}원
공급가+부가세: {total_invoice:,.0f}원

제조원가: {total_cost:,.0f}원
물류이익: {total_profit:,.0f}원
물류 이익률: {total_margin:.1%}
""")

    for _, row in store_report.iterrows():

        lines.append(f"""
━━━━━━━━━━
[{row['거래처명']}]

공급가액: {row['공급가액']:,.0f}원
부가세: {row['부가세']:,.0f}원
청구합계: {row['청구합계']:,.0f}원

총제조원가: {row['총제조원가']:,.0f}원
물류이익: {row['물류이익']:,.0f}원
물류이익률: {row['물류이익률']:.1%}
""")

    if len(missing) > 0:
        lines.append("\n⚠ 원가 미등록 SKU")
        for _, row in missing.iterrows():
            lines.append(f"- {row['품목코드']} / {row['품목명(규격)']}")

    sku_report["공급비중"] = (
        sku_report["공급가액"] /
        (total_supply if total_supply != 0 else 1)
    )

    sku_report["이익비중"] = (
        sku_report["물류이익"] /
        (total_profit if total_profit != 0 else 1)
    )
    
    sku_report = sku_report.sort_values(
    "공급가액",
    ascending=False
)
    store_sku_report = (
        erp.groupby(["거래처명", "품목코드", "품목명(규격)"])
        .agg(
            수량=("수량", "sum"),
            공급가액=("공급가액", "sum"),
            총제조원가=("총제조원가", "sum"),
            물류이익=("물류이익", "sum"),
        )
        .reset_index()
    )

    store_sku_report["물류이익률"] = (
        store_sku_report["물류이익"] /
        store_sku_report["공급가액"].replace(0, pd.NA)
    )

    store_sku_report = store_sku_report.sort_values(
        ["거래처명", "공급가액"],
        ascending=[True, False]
    )

    store_sku_report["_sort"] = store_sku_report["거래처명"].map(STORE_ORDER).fillna(99)
    store_sku_report = (
        store_sku_report
        .sort_values(["_sort", "거래처명", "공급가액"], ascending=[True, True, False])
        .drop(columns=["_sort"])
        .reset_index(drop=True)
    )
    
    with pd.ExcelWriter(
        "월말_물류리포트.xlsx",
        engine="openpyxl"
    ) as writer:

        # 1) SUMMARY 시트
        summary_df = pd.DataFrame({
            "구분": [
                "전체 공급가액",
                "전체 부가세",
                "전체 청구합계",
                "전체 제조원가",
                "전체 물류이익",
                "전체 물류이익률",
            ],
            "값": [
                total_supply,
                total_vat,
                total_invoice,
                total_cost,
                total_profit,
                total_margin,
            ]
        })

        summary_df.to_excel(
            writer,
            sheet_name="SUMMARY",
            index=False,
            startrow=2
        )

        store_report.to_excel(
            writer,
            sheet_name="STORE_REPORT",
            index=False
        )

        sku_report.to_excel(
            writer,
            sheet_name="SKU_REPORT",
            index=False
        )
        
        store_sku_report.to_excel(
            writer,
            sheet_name="STORE_SKU_REPORT",
            index=False
        )

        workbook = writer.book

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference

        navy_fill = PatternFill("solid", fgColor="1F4E78")
        dark_fill = PatternFill("solid", fgColor="1F2933")
        green_fill = PatternFill("solid", fgColor="1F7A68")
        green_soft_fill = PatternFill("solid", fgColor="DDEFEA")
        yellow_fill = PatternFill("solid", fgColor="FFF7D6")
        light_fill = PatternFill("solid", fgColor="D9EAF7")
        gray_fill = PatternFill("solid", fgColor="F2F2F2")
        red_font = Font(color="C00000", bold=True)
        white_bold = Font(color="FFFFFF", bold=True)
        black_bold = Font(color="000000", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        def style_sheet(ws):
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
                    cell.border = thin_border

            for cell in ws[1]:
                cell.fill = navy_fill
                cell.font = white_bold

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))

                    if isinstance(cell.value, (int, float)):
                        if "율" in str(ws.cell(row=1, column=cell.column).value) or "비중" in str(ws.cell(row=1, column=cell.column).value):
                            cell.number_format = "0.0%"
                        else:
                            cell.number_format = '#,##0'

                ws.column_dimensions[col_letter].width = min(max_length + 4, 35)

        # SUMMARY 디자인
        ws = workbook["SUMMARY"]
        ws["A1"] = "월말 물류이익 리포트"
        ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
        ws.merge_cells("A1:B1")

        for cell in ws[3]:
            cell.fill = navy_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=4, max_row=9):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        ws["B4"].number_format='#,##0"원"'
        ws["B5"].number_format='#,##0"원"'
        ws["B6"].number_format='#,##0"원"'
        ws["B7"].number_format='#,##0"원"'
        ws["B8"].number_format='#,##0"원"'
        ws["B9"].number_format='0.0%'

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22

        # SUMMARY KPI cards
        kpis = [
            ("공급가액", total_supply, '#,##0"원"', "D3:E4"),
            ("제조원가", total_cost, '#,##0"원"', "F3:G4"),
            ("물류이익", total_profit, '#,##0"원"', "D6:E7"),
            ("물류이익률", total_margin, "0.0%", "F6:G7"),
        ]
        for label, value, number_format, cell_range in kpis:
            ws.merge_cells(cell_range)
            card = ws[cell_range.split(":")[0]]
            card.value = f"{label}\n{value}"
            card.fill = green_soft_fill if label != "물류이익률" else yellow_fill
            card.font = Font(size=13, bold=True, color="1F2933")
            card.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            card.border = thin_border
            card.number_format = number_format
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        for row in [3, 4, 6, 7]:
            ws.row_dimensions[row].height = 26

        # STORE_REPORT 디자인
        ws = workbook["STORE_REPORT"]
        style_sheet(ws)

        for row in range(2, ws.max_row + 1):
            margin_cell = ws.cell(row=row, column=7)
            margin_cell.number_format = "0.0%"

            if margin_cell.value is not None and margin_cell.value >= 0.4:
                margin_cell.fill = light_fill
                margin_cell.font = black_bold

        # SKU_REPORT 디자인
        ws = workbook["SKU_REPORT"]
        style_sheet(ws)

        for row in range(2, ws.max_row + 1):
            profit_cell = ws.cell(row=row, column=6)
            margin_cell = ws.cell(row=row, column=7)
            supply_ratio_cell = ws.cell(row=row, column=8)
            profit_ratio_cell = ws.cell(row=row, column=9)

            margin_cell.number_format = "0.0%"
            supply_ratio_cell.number_format = "0.0%"
            profit_ratio_cell.number_format = "0.0%"
            ws.cell(row=row, column=3).number_format = '#,##0"개"'
            ws.cell(row=row, column=4).number_format = '#,##0"원"'
            ws.cell(row=row, column=5).number_format = '#,##0"원"'
            ws.cell(row=row, column=6).number_format = '#,##0"원"'
            
            if profit_cell.value is not None and profit_cell.value < 0:
                profit_cell.font = red_font

        # STORE_SKU_REPORT 디자인
        ws = workbook["STORE_SKU_REPORT"]
        style_sheet(ws)

        for row in range(2, ws.max_row + 1):

            profit_cell = ws.cell(row=row, column=7)
            margin_cell = ws.cell(row=row, column=8)

            margin_cell.number_format = "0.0%"

            ws.cell(row=row, column=4).number_format = '#,##0"개"'
            ws.cell(row=row, column=5).number_format = '#,##0"원"'
            ws.cell(row=row, column=6).number_format = '#,##0"원"'
            ws.cell(row=row, column=7).number_format = '#,##0"원"'

            if profit_cell.value is not None and profit_cell.value < 0:
                profit_cell.font = red_font

        def apply_report_layout(ws, title_text, money_columns=(), percent_columns=(), qty_columns=()):
            ws.insert_rows(1)
            max_col = ws.max_column
            max_row = ws.max_row
            title_range = f"A1:{get_column_letter(max_col)}1"
            ws.merge_cells(title_range)
            ws["A1"] = title_text
            ws["A1"].fill = dark_fill
            ws["A1"].font = Font(size=15, bold=True, color="FFFFFF")
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 28

            header = ws[2]
            for cell in header:
                cell.fill = green_fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row in ws.iter_rows(min_row=3, max_row=max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    if cell.row % 2 == 1:
                        cell.fill = PatternFill("solid", fgColor="FAFAFA")

            for col in money_columns:
                for cell in ws[get_column_letter(col)][2:]:
                    cell.number_format = '#,##0"원"'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for col in percent_columns:
                for cell in ws[get_column_letter(col)][2:]:
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for col in qty_columns:
                for cell in ws[get_column_letter(col)][2:]:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col_letter].width = min(max(max_length + 3, 11), 34)

            ws.freeze_panes = "A3"
            ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"

        apply_report_layout(
            workbook["STORE_REPORT"],
            "지점별 물류이익 요약",
            money_columns=(2, 3, 4, 5, 6),
            percent_columns=(7,),
        )
        apply_report_layout(
            workbook["SKU_REPORT"],
            "SKU별 공급/이익 분석",
            money_columns=(4, 5, 6),
            percent_columns=(7, 8, 9),
            qty_columns=(3,),
        )
        apply_report_layout(
            workbook["STORE_SKU_REPORT"],
            "지점별 SKU 상세",
            money_columns=(5, 6, 7),
            percent_columns=(8,),
            qty_columns=(4,),
        )

        def style_block(ws, cell_range, fill, font=None):
            for row in ws[cell_range]:
                for cell in row:
                    cell.fill = fill
                    cell.border = thin_border
                    if font:
                        cell.font = font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        def create_ceo_dashboard():
            ws = workbook.create_sheet("CEO_DASHBOARD", 0)
            ws.sheet_view.showGridLines = False

            ws.merge_cells("A1:L1")
            ws["A1"] = "월말 물류 CEO 대시보드"
            ws["A1"].fill = dark_fill
            ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 34

            ws.merge_cells("A2:L2")
            ws["A2"] = "지점별 공급가액, 제조원가, 물류이익 흐름을 한 화면에서 확인합니다."
            ws["A2"].fill = gray_fill
            ws["A2"].font = Font(size=11, color="5D6975")
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

            cards = [
                ("공급가액", total_supply, '#,##0"원"', "A4:C6", green_soft_fill),
                ("제조원가", total_cost, '#,##0"원"', "D4:F6", light_fill),
                ("물류이익", total_profit, '#,##0"원"', "G4:I6", yellow_fill),
                ("물류이익률", total_margin, "0.0%", "J4:L6", green_soft_fill),
            ]
            for label, value, number_format, cell_range, fill in cards:
                ws.merge_cells(cell_range)
                cell = ws[cell_range.split(":")[0]]
                cell.value = value
                cell.number_format = number_format
                cell.fill = fill
                cell.font = Font(size=18, bold=True, color="1F2933")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                label_cell = ws.cell(row=cell.row, column=cell.column)
                ws.cell(row=cell.row - 1, column=cell.column).value = label
                ws.cell(row=cell.row - 1, column=cell.column).font = Font(size=11, bold=True, color="1F2933")

            ws["A8"] = "지점별 성과"
            ws["A8"].font = Font(size=14, bold=True, color="1F2933")
            store_headers = ["지점", "공급가액", "제조원가", "물류이익", "이익률"]
            for col, header in enumerate(store_headers, 1):
                cell = ws.cell(row=9, column=col, value=header)
                cell.fill = green_fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for r, (_, row) in enumerate(store_report.iterrows(), 10):
                values = [
                    row["거래처명"],
                    row["공급가액"],
                    row["총제조원가"],
                    row["물류이익"],
                    row["물류이익률"],
                ]
                for c, value in enumerate(values, 1):
                    cell = ws.cell(row=r, column=c, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right" if c > 1 else "center")
                    if r % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="FAFAFA")
                    if c in (2, 3, 4):
                        cell.number_format = '#,##0"원"'
                    if c == 5:
                        cell.number_format = "0.0%"

            chart = BarChart()
            chart.title = "지점별 공급가액"
            chart.y_axis.title = "공급가액"
            chart.x_axis.title = "지점"
            data = Reference(ws, min_col=2, min_row=9, max_row=9 + len(store_report))
            cats = Reference(ws, min_col=1, min_row=10, max_row=9 + len(store_report))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.legend = None
            chart.height = 8
            chart.width = 16
            ws.add_chart(chart, "G8")

            ws["A18"] = "Top SKU 공급가액"
            ws["A18"].font = Font(size=14, bold=True, color="1F2933")
            sku_headers = ["품목코드", "품목명", "공급가액", "물류이익", "이익률"]
            for col, header in enumerate(sku_headers, 1):
                cell = ws.cell(row=19, column=col, value=header)
                cell.fill = green_fill
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            top_sku = sku_report.sort_values("공급가액", ascending=False).head(8)
            for r, (_, row) in enumerate(top_sku.iterrows(), 20):
                values = [
                    row["품목코드"],
                    row["품목명(규격)"],
                    row["공급가액"],
                    row["물류이익"],
                    row["물류이익률"],
                ]
                for c, value in enumerate(values, 1):
                    cell = ws.cell(row=r, column=c, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right" if c >= 3 else "left")
                    if r % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="FAFAFA")
                    if c in (3, 4):
                        cell.number_format = '#,##0"원"'
                    if c == 5:
                        cell.number_format = "0.0%"

            ws["G18"] = "CEO 체크 포인트"
            ws["G18"].font = Font(size=14, bold=True, color="1F2933")
            notes = [
                f"전체 물류이익률: {total_margin:.1%}",
                f"최고 공급 지점: {store_report.sort_values('공급가액', ascending=False).iloc[0]['거래처명']}",
                f"최고 이익 SKU: {sku_report.sort_values('물류이익', ascending=False).iloc[0]['품목명(규격)']}",
            ]
            for idx, note in enumerate(notes, 19):
                ws.merge_cells(start_row=idx, start_column=7, end_row=idx, end_column=12)
                cell = ws.cell(row=idx, column=7, value=note)
                cell.fill = green_soft_fill
                cell.border = thin_border
                cell.font = Font(size=11, color="1F2933")
                cell.alignment = Alignment(horizontal="left", vertical="center")

            widths = {
                "A": 14, "B": 34, "C": 16, "D": 16, "E": 12,
                "F": 3, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            for row in range(1, 30):
                ws.row_dimensions[row].height = 22

            ws.freeze_panes = "A8"

        create_ceo_dashboard()

        # 보기 편하게 시트 순서
        workbook._sheets = [
            workbook["CEO_DASHBOARD"],
            workbook["SUMMARY"],
            workbook["STORE_REPORT"],
            workbook["SKU_REPORT"],
            workbook["STORE_SKU_REPORT"],
        ]

    return "\n".join(lines)


if __name__ == "__main__":

    report = run_monthly_logistics_report()

    print(report)

    if not TELEGRAM_BOT_TOKEN:
        raise Exception("TELEGRAM_BOT_TOKEN 환경변수 없음")

    telegram_url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"

    MAX_LEN = 3500

    for chat_id in TELEGRAM_CHAT_IDS:

        chunks = []
        sections = report.split("━━━━━━━━━━")
        current = ""

        for section in sections:

            section = section.strip()

            if not section:
                continue

            section = "━━━━━━━━━━\n" + section

            if len(current) + len(section) > MAX_LEN:

                if current:
                    chunks.append(current)

                current = section

            else:

                if current:
                    current += "\n" + section
                else:
                    current = section

        if current:
            chunks.append(current)

        for chunk in chunks:

            payload = {
                "chat_id": chat_id,
                "text": chunk
            }

            res = requests.post(
                telegram_url,
                data=payload,
                timeout=15
            )

            print(
                "텔레그램 응답:",
                chat_id,
                res.status_code,
                res.text
            )

    for chat_id in TELEGRAM_CHAT_IDS:

        with open(
            "월말_물류리포트.xlsx",
            "rb"
        ) as file:

            res = requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument",
                data={
                    "chat_id": chat_id
                },
                files={
                    "document": file
                }
            )

            print(
                "엑셀 전송:",
                chat_id,
                res.status_code,
                res.text
            )

    print("텔레그램 전송 완료")
    
